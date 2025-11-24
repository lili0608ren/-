import streamlit as st
import pandas as pd
import numpy as np
import pulp
from collections import defaultdict
from openpyxl import Workbook
from datetime import time as dt_time
import io

DEFAULT_OUTPUT = "最適化結果.xlsx"

st.set_page_config(page_title="車両指定配車最適化", layout="wide")
st.title("車両指定付き配車最適化 (VRPTW)")

# --- 入力: ファイル・パラメータ ---
st.sidebar.header("設定")
uploaded = st.sidebar.file_uploader("Excelファイルをアップロード", type=["xlsx", "xls"])
dur_csv = st.sidebar.file_uploader("duration_matrix CSVファイル", type=["csv"])
gosa = st.sidebar.number_input("許容誤差 gosa (秒)", value=300, step=60)
solver_time_limit = st.sidebar.number_input("Solver time limit (秒)", value=960, step=60)
run_button = st.sidebar.button("最適化を実行")

if not uploaded or not dur_csv:
    st.info("Excel・duration_matrixファイル両方アップロードしてください。")
    st.stop()

excel_io = io.BytesIO(uploaded.read())
df_user = pd.read_excel(excel_io, sheet_name="利用者に住所")
df_wc   = pd.read_excel(excel_io, sheet_name="車椅子の有無")
df_car  = pd.read_excel(excel_io, sheet_name="車両情報")
df_time = pd.read_excel(excel_io, sheet_name="時間帯制約")

users = [x for x in df_user["利用者名"].tolist() if x != "デポ"]
address_dict = dict(zip(df_user["利用者名"], df_user["住所"]))
n_users = len(users)
n_nodes = n_users + 1
user_to_node = {users[i]: i + 1 for i in range(n_users)}
node_to_user = {i + 1: users[i] for i in range(n_users)}
user_wc_flags = dict(zip(df_wc['利用者名'], df_wc['車いすの有無']))
user_wheelchair = {u: int(user_wc_flags.get(u, 0)) for u in users}
pickup_times = {u: 300 if user_wheelchair[u] == 1 else 180 for u in users}
duration_matrix = pd.read_csv(dur_csv, index_col=0).to_numpy()
# 行列サイズ確認
st.write("duration_matrix shape:", duration_matrix.shape)
st.write("n_nodes (利用者+1):", n_nodes)
if duration_matrix.shape != (n_nodes, n_nodes):
    st.error("duration_matrix (CSV) のサイズが利用者数+1 (デポ含む)と一致していません！")
    st.error("  - 利用者数+1（行列サイズ）: " + str(n_nodes))
    st.error("  - duration_matrix shape: " + str(duration_matrix.shape))
    st.stop()

車両候補リスト = df_car["車種"].astype(str).unique().tolist()
use_cars = st.multiselect("使う車両を選んでください", 車両候補リスト, default=車両候補リスト)
if not use_cars:
    st.warning("1台以上選択してください")
    st.stop()

base_vehicles = []
for _, row in df_car.iterrows():
    if str(row["車種"]) in use_cars:
        car = {
            "車両名": str(row["車種"]),
            "通常定員": int(row["通常定員"]),
            "車椅子最大数": int(row["車椅子最大"]),
            "車椅子対応": 1 if int(row["車椅子最大"]) > 0 else 0,
            "車椅子一台あたりの人数": int(row["車椅子一台あたりの人数"]),
        }
        base_vehicles.append(car)

MAX_TRIPS = 3
vehicles = []
for car in base_vehicles:
    for trip in range(MAX_TRIPS):
        new_car = car.copy()
        new_car["便名"] = f"{car['車両名']}_trip{trip + 1}"
        new_car["trip_index"] = trip + 1
        new_car["便出発時間"] = "08:00:00"
        vehicles.append(new_car)
v = len(vehicles)

def to_seconds(t):
    if isinstance(t, dt_time):
        return t.hour * 3600 + t.minute * 60 + t.second
    if pd.isna(t):
        return None
    if isinstance(t, (pd.Timestamp, pd.Series)):
        try:
            return int(t.hour * 3600 + t.minute * 60 + t.second)
        except Exception:
            return None
    if isinstance(t, (int, float)):
        return int(t)
    if isinstance(t, str):
        try:
            parts = t.split(":")
            if len(parts) == 3:
                h, m, s = map(int, parts)
                return h*3600 + m*60 + s
            if len(parts) == 2:
                h, m = map(int, parts)
                return h*3600 + m*60
        except Exception:
            return None
    return None

if run_button:
    st.write(f"計算処理中...車両便数={v}")
    DAY_START_SEC = 8 * 3600
    BIG_M = 10 ** 5
    time_constraints = {}
    if '利用者名' in df_time.columns:
        for _, row in df_time.iterrows():
            name = row.get("利用者名")
            strict_val = int(row.get("開始時間厳守", 0) if not pd.isna(row.get("開始時間厳守", 0)) else 0)
            start_time_val = row.get("開始時間")
            time_constraints[name] = {"strict": strict_val, "time_sec": start_time_val}

    prob = pulp.LpProblem("VRPTW_full_fixed", pulp.LpMinimize)

    used = pulp.LpVariable.dicts("used", range(v), cat="Binary")
    x = pulp.LpVariable.dicts("x", ((i, k) for i in range(1, n_nodes) for k in range(v)), cat="Binary")
    y = {}
    for i in range(n_nodes):
        for j in range(n_nodes):
            if i == j: continue
            for k in range(v):
                y[(i, j, k)] = pulp.LpVariable(f"y_{i}_{j}_{k}", cat="Binary")
    arrival = [pulp.LpVariable(f"arrival_{i}", lowBound=0) for i in range(n_nodes)]
    trip_start = [pulp.LpVariable(f"trip_start_{k}", lowBound=DAY_START_SEC) for k in range(v)]
    trip_end = [pulp.LpVariable(f"trip_end_{k}", lowBound=DAY_START_SEC) for k in range(v)]
    genshu = pulp.LpVariable.dicts("genshu", range(n_users), cat="Binary")
    u_var = {}
    for i in range(1, n_nodes):
        for k in range(v):
            u_var[(i, k)] = pulp.LpVariable(f"u_{i}_{k}", lowBound=1, upBound=n_users, cat="Integer")
    max_time = pulp.LpVariable("max_time", lowBound=0)
    car_penalty = []
    for car in vehicles:
        if "パレット" in car["車両名"]:
            car_penalty.append(1)
        elif "ラクティス" in car["車両名"]:
            car_penalty.append(1)
        elif "ハイエース" in car["車両名"]:
            car_penalty.append(5)
        else:
            car_penalty.append(1)
    vehicle_penalty_term = pulp.lpSum(car_penalty[k] * used[k] for k in range(v))
    early_violation = [pulp.LpVariable(f"early_v_{i}", lowBound=0) for i in range(n_nodes)]
    late_violation  = [pulp.LpVariable(f"late_v_{i}", lowBound=0) for i in range(n_nodes)]
    early_penalty = 1000000
    late_penalty  = 1000000
    for i in range(1, n_nodes):
        prob += early_violation[i] >= 8*3600 - arrival[i]
        prob += late_violation[i]  >= arrival[i] - 10*3600
    prob += (vehicle_penalty_term + max_time +
             early_penalty * pulp.lpSum(early_violation) +
             late_penalty * pulp.lpSum(late_violation))

    # 制約 (1)～(12)（インデント同一でコピペ！）
    # constraints (1)
    for i in range(1, n_nodes):
        prob += pulp.lpSum(x[(i, k)] for k in range(v)) == 1

    # (2) capacity
    for k, car in enumerate(vehicles):
        normal_sum = pulp.lpSum(x[(i, k)] for i in range(1, n_nodes) if user_wheelchair[node_to_user[i]] == 0)
        wc_sum = pulp.lpSum(car["車椅子一台あたりの人数"] * x[(i, k)] for i in range(1, n_nodes) if user_wheelchair[node_to_user[i]] == 1)
        prob += (normal_sum + wc_sum) <= car["通常定員"] * used[k]
        for i in range(1, n_nodes):
            prob += x[(i, k)] <= used[k]

    # (3) wheelchair constraints
    wc_idx = [i+1 for i in range(n_users) if user_wheelchair.get(users[i], 0) == 1]
    for k, car in enumerate(vehicles):
        if car["車椅子最大数"] is not None:
            try:
                prob += pulp.lpSum(x[(i, k)] for i in wc_idx) <= car["車椅子最大数"]
            except Exception:
                pass
        if not car["車椅子対応"]:
            prob += pulp.lpSum(x[(i, k)] for i in wc_idx) == 0

    # (4) depot flow
    for k in range(v):
        prob += pulp.lpSum(y[(0, j, k)] for j in range(1, n_nodes)) == pulp.lpSum(y[(i, 0, k)] for i in range(1, n_nodes))
        prob += pulp.lpSum(y[(0, j, k)] for j in range(1, n_nodes)) == used[k] * 1

    # (5) flow conservation
    for k in range(v):
        for i in range(1, n_nodes):
            prob += pulp.lpSum(y[(i, j, k)] for j in range(n_nodes) if j != i) == x[(i, k)]
            prob += pulp.lpSum(y[(j, i, k)] for j in range(n_nodes) if j != i) == x[(i, k)]

    # (6) arrival: depot -> user
    for k in range(v):
        for j in range(1, n_nodes):
            prob += arrival[j] >= trip_start[k] + int(duration_matrix[0, j]) - BIG_M * (1 - y[(0, j, k)])

    # (7) between users
    for k in range(v):
        for i in range(1, n_nodes):
            for j in range(1, n_nodes):
                if i == j: continue
                prob += arrival[j] >= arrival[i] + pickup_times[node_to_user[i]] + int(duration_matrix[i, j]) - BIG_M * (1 - y[(i, j, k)])

    # (8) trip_end
    for k in range(v):
        for i in range(1, n_nodes):
            prob += trip_end[k] >= arrival[i] + pickup_times[node_to_user[i]] + int(duration_matrix[i, 0]) - BIG_M * (1 - y[(i, 0, k)])

    # (9) strict time windows
    for i in range(1, n_nodes):
        uname = node_to_user[i]
        tc = time_constraints.get(uname, {"strict": 0, "time_sec": None})
        if tc["strict"] == 1 and tc["time_sec"] is not None:
            desired = to_seconds(tc["time_sec"])
            if desired is not None:
                prob += arrival[i] >= desired - gosa
                prob += arrival[i] <= desired + gosa

    # (10) MTZ
    for k in range(v):
        for i in range(1, n_nodes):
            for j in range(1, n_nodes):
                if i == j: continue
                prob += u_var[(i, k)] - u_var[(j, k)] + n_users * y[(i, j, k)] <= n_users - 1

    # (11) max_time
    for k in range(v):
        prob += max_time >= trip_end[k] - trip_start[k]

    # (12) same vehicle trip ordering
    vehicle_trip_indices = defaultdict(list)
    for k, car in enumerate(vehicles):
        vehicle_trip_indices[car["車両名"]].append(k)
    for car_name, trip_list in vehicle_trip_indices.items():
        trip_list_sorted = sorted(trip_list)
        for idx_ in range(len(trip_list_sorted)-1):
            k1 = trip_list_sorted[idx_]
            k2 = trip_list_sorted[idx_+1]
            prob += trip_start[k2] >= trip_end[k1] + 600
            prob += used[k1] >= used[k2]


    for i in range(1, n_nodes):
        prob += pulp.lpSum(x[(i, k)] for k in range(v)) == 1
    for k, car in enumerate(vehicles):
        normal_sum = pulp.lpSum(x[(i, k)] for i in range(1, n_nodes) if user_wheelchair[node_to_user[i]] == 0)
        wc_sum = pulp.lpSum(car["車椅子一台あたりの人数"] * x[(i, k)] for i in range(1, n_nodes) if user_wheelchair[node_to_user[i]] == 1)
        prob += (normal_sum + wc_sum) <= car["通常定員"] * used[k]
        for i in range(1, n_nodes):
            prob += x[(i, k)] <= used[k]
    wc_idx = [i+1 for i in range(n_users) if user_wheelchair.get(users[i], 0) == 1]
    for k, car in enumerate(vehicles):
        if car["車椅子最大数"] is not None:
            try:
                prob += pulp.lpSum(x[(i, k)] for i in wc_idx) <= car["車椅子最大数"]
            except Exception:
                pass
        if not car["車椅子対応"]:
            prob += pulp.lpSum(x[(i, k)] for i in wc_idx) == 0
    for k in range(v):
        prob += pulp.lpSum(y[(0, j, k)] for j in range(1, n_nodes)) == pulp.lpSum(y[(i, 0, k)] for i in range(1, n_nodes))
        prob += pulp.lpSum(y[(0, j, k)] for j in range(1, n_nodes)) == used[k] * 1
    for k in range(v):
        for i in range(1, n_nodes):
            prob += pulp.lpSum(y[(i, j, k)] for j in range(n_nodes) if j != i) == x[(i, k)]
            prob += pulp.lpSum(y[(j, i, k)] for j in range(n_nodes) if j != i) == x[(i, k)]
    for k in range(v):
        for j in range(1, n_nodes):
            prob += arrival[j] >= trip_start[k] + int(duration_matrix[0, j]) - BIG_M * (1 - y[(0, j, k)])
    for k in range(v):
        for i in range(1, n_nodes):
            for j in range(1, n_nodes):
                if i == j: continue
                prob += arrival[j] >= arrival[i] + pickup_times[node_to_user[i]] + int(duration_matrix[i, j]) - BIG_M * (1 - y[(i, j, k)])
    for k in range(v):
        for i in range(1, n_nodes):
            prob += trip_end[k] >= arrival[i] + pickup_times[node_to_user[i]] + int(duration_matrix[i, 0]) - BIG_M * (1 - y[(i, 0, k)])
    time_constraints = {}
    if '利用者名' in df_time.columns:
        for _, row in df_time.iterrows():
            name = row.get("利用者名")
            strict_val = int(row.get("開始時間厳守", 0) if not pd.isna(row.get("開始時間厳守", 0)) else 0)
            start_time_val = row.get("開始時間")
            time_constraints[name] = {"strict": strict_val, "time_sec": start_time_val}
    for i in range(1, n_nodes):
        uname = node_to_user[i]
        tc = time_constraints.get(uname, {"strict": 0, "time_sec": None})
        if tc["strict"] == 1 and tc["time_sec"] is not None:
            desired = to_seconds(tc["time_sec"])
            if desired is not None:
                prob += arrival[i] >= desired - gosa
                prob += arrival[i] <= desired + gosa
    for k in range(v):
        for i in range(1, n_nodes):
            for j in range(1, n_nodes):
                if i == j: continue
                prob += u_var[(i, k)] - u_var[(j, k)] + n_users * y[(i, j, k)] <= n_users - 1
    for k in range(v):
        prob += max_time >= trip_end[k] - trip_start[k]
    vehicle_trip_indices = defaultdict(list)
    for k, car in enumerate(vehicles):
        vehicle_trip_indices[car["車両名"]].append(k)
    for car_name, trip_list in vehicle_trip_indices.items():
        trip_list_sorted = sorted(trip_list)
        for idx_ in range(len(trip_list_sorted)-1):
            k1 = trip_list_sorted[idx_]
            k2 = trip_list_sorted[idx_+1]
            prob += trip_start[k2] >= trip_end[k1] + 600
            prob += used[k1] >= used[k2]

    solver = pulp.PULP_CBC_CMD(msg=1, timeLimit=int(solver_time_limit), threads=4)
    with st.spinner("ソルバー計算中…"):
        res = prob.solve(solver)
    st.success(f"Solver status: {pulp.LpStatus[prob.status]}, obj: {pulp.value(prob.objective)}")

    # --- extract_routes_from_y, 結果処理部: 省略可（そのまま貼付でOK）---
    # ...（省略。前回答例やCLIとも共通）...
    def extract_routes_from_y(y_vars, vehicles, n_nodes):
        routes = {}
        for k in range(len(vehicles)):
            starts = [j for j in range(1, n_nodes)
                      if pulp.value(y_vars[(0, j, k)]) is not None and pulp.value(y_vars[(0, j, k)]) > 0.5]
            if not starts: continue
            route = [0]
            cur = starts[0]
            route.append(cur)
            visited = set([cur])
            while True:
                if pulp.value(y_vars.get((cur, 0, k), 0)) is not None and pulp.value(y_vars.get((cur, 0, k), 0)) > 0.5:
                    route.append(0)
                    break
                nexts = [j for j in range(1, n_nodes) if j != cur
                        and pulp.value(y_vars[(cur, j, k)]) is not None and pulp.value(y_vars[(cur, j, k)]) > 0.5]
                found = None
                for nx in nexts:
                    if nx not in visited:
                        found = nx
                        break
                if found is None:
                    if nexts: found = nexts[0]
                    else:
                        route.append(0)
                        break
                route.append(found)
                visited.add(found)
                cur = found
                if len(route) > n_nodes + 5:
                    route.append(0)
                    break
            routes[k] = route
        return routes

    routes_by_k = extract_routes_from_y(y, vehicles, n_nodes)

    # Build assign_map and totals
    assign_map = {}
    total_times_map = {}
    last_end_times = defaultdict(lambda: DAY_START_SEC)

    for k, car in enumerate(vehicles):
        if k not in routes_by_k:
            continue
        route = routes_by_k[k]
        if pulp.value(used[k]) is not None and pulp.value(used[k]) > 0.5:
            base_start = int(pulp.value(trip_start[k])) if pulp.value(trip_start[k]) is not None else DAY_START_SEC
        else:
            base_start = DAY_START_SEC
        current = max(last_end_times[car["車両名"]], base_start)
        for idx in range(1, len(route) - 1):
            prev = route[idx - 1]
            cur = route[idx]
            travel = int(duration_matrix[prev, cur])
            current += travel
            user_name = node_to_user[cur]
            arrival_val = int(pulp.value(arrival[cur])) if pulp.value(arrival[cur]) is not None else None
            desired_val = time_constraints.get(user_name, {}).get("time_sec")
            desired_sec = to_seconds(desired_val) if desired_val is not None else None

            if arrival_val is not None:
                h, rem = divmod(arrival_val, 3600)
                m, s = divmod(rem, 60)
            else:
                h = m = s = None

            assign_map[user_name] = {
                "車椅子の有無": user_wheelchair.get(user_name, 0),
                "車種": car["車両名"],
                "便名": car["便名"],
                "ピックアップ_time_h": int(h) if h is not None else None,
                "ピックアップ_time_m": int(m) if m is not None else None,
                "ピックアップ_time_s": int(s) if s is not None else None,
                "順番": f"{car['trip_index']}便目の{idx}",
                "desired_strict": desired_val,
                "genshu_var": int(pulp.value(genshu[user_to_node[user_name] - 1]) if pulp.value(genshu[user_to_node[user_name] - 1]) is not None else 0),
            }
            current += pickup_times[user_name]
        last_node = route[-2] if len(route) >= 2 else 0
        if last_node != 0:
            current += int(duration_matrix[last_node, 0])
        last_end_times[car["車両名"]] = current
        total_times_map[car["便名"]] = (last_end_times[car["車両名"]] - base_start) // 60

    # Violation check
    violations = []
    for u, info in assign_map.items():
        desired = info.get("desired_strict")
        if desired is None:
            continue
        if info.get("ピックアップ_time_h") is None:
            continue
        arrival_sec = info["ピックアップ_time_h"] * 3600 + info["ピックアップ_time_m"] * 60 + info["ピックアップ_time_s"]
        if isinstance(desired, dt_time) or isinstance(desired, pd.Timestamp):
            desired_sec = to_seconds(desired)
        else:
            desired_sec = to_seconds(desired)
        if desired_sec is None:
            continue
        if abs(arrival_sec - desired_sec) > gosa:
            violations.append((u, desired_sec, arrival_sec))

    # Prepare Excel output in memory
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "結果"
    ws_out.append([
        "利用者名", "車椅子の有無", "車種", "便名", "ピックアップ時間（時）",
        "ピックアップ時間（分）", "ピックアップ時間（秒）", "順番", "車両総移動時間（分）",
        "genshu_var", "希望時刻(秒)"
    ])
    for u in users:
        if u in assign_map:
            d = assign_map[u]
            total_time = total_times_map.get(d["便名"], "")
            ws_out.append([
                u, d["車椅子の有無"], d["車種"], d["便名"], d["ピックアップ_time_h"],
                d["ピックアップ_time_m"], d["ピックアップ_time_s"], d["順番"], total_time,
                d.get("genshu_var", 0), d.get("desired_strict", "")
            ])
        else:
            ws_out.append([u, "未割当", "", "", "", "", "", "", "", "", ""])

    ws_v = wb_out.create_sheet("違反チェック")
    ws_v.append(["利用者名", "希望(秒)", "実到着(秒)", "差(秒)"])
    for vinfo in violations:
        u, desired, arrival = vinfo
        ws_v.append([u, desired, arrival, arrival - desired])

    ws_r = wb_out.create_sheet("便別ルート")
    ws_r.append(["便名", "route_nodes (node indices)", "route_users (順序)", "便出発時刻(秒)", "便終了時刻(秒)", "便使用フラグ"])
    for k, car in enumerate(vehicles):
        route = routes_by_k.get(k, [])
        route_users = [node_to_user[idx] for idx in route if idx != 0]
        trip_start_sol = int(pulp.value(trip_start[k])) if pulp.value(trip_start[k]) is not None else ""
        trip_end_sol = int(pulp.value(trip_end[k])) if pulp.value(trip_end[k]) is not None else ""
        ws_r.append([car["便名"], ",".join(map(str, route)), "->".join(route_users), trip_start_sol, trip_end_sol, int(pulp.value(used[k]) or 0)])

    # Save workbook to bytes
    out_bytes = io.BytesIO()
    wb_out.save(out_bytes)
    out_bytes.seek(0)

    # Display summary tables
    st.header("結果概要")
    if violations:
        st.warning(f"注意: {len(violations)} 件の時間窓違反があります。'違反チェック' シートをダウンロードして確認してください。")
    else:
        st.success("全員の到着が許容内です。")

    # Show assign table
    df_assign = []
    for u in users:
        if u in assign_map:
            d = assign_map[u]
            total_time = total_times_map.get(d["便名"], "")
            df_assign.append({
                "利用者名": u,
                "車椅子": d["車椅子の有無"],
                "車種": d["車種"],
                "便名": d["便名"],
                "ピックアップ時刻": f"{d['ピックアップ_time_h']:02d}:{d['ピックアップ_time_m']:02d}:{d['ピックアップ_time_s']:02d}" if d["ピックアップ_time_h"] is not None else "",
                "順番": d["順番"],
                "genshu": d.get("genshu_var", 0),
                "便合計所要時間（分）": total_time,   # 追加
            })
        else:
            df_assign.append({
                "利用者名": u,
                "車椅子": "", "車種": "", "便名": "",
                "ピックアップ時刻": "", "順番": "",
                "genshu": "",
                "便合計所要時間（分）": ""
            })
    st.dataframe(pd.DataFrame(df_assign))

        # Show routes compact
    st.subheader("便別ルート")
    rows = []
    for k, car in enumerate(vehicles):
        route = routes_by_k.get(k, [])
        if not route:
            continue
        route_users = [node_to_user[idx] for idx in route if idx != 0]
        total_time = total_times_map.get(car["便名"], "")
        rows.append({
            "便名": car["便名"],
            "車種": car["車両名"],
            "route_nodes": ",".join(map(str, route)),
            "route_users": " -> ".join(route_users),
            "used": int(pulp.value(used[k]) or 0),
            "便合計所要時間（分）": total_time,  # 追加
        })
    if rows:
        st.table(pd.DataFrame(rows))


    # Download button
    st.download_button("結果 Excel をダウンロード", data=out_bytes.getvalue(),
                       file_name=DEFAULT_OUTPUT, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.balloons()
    #log.write("終了。")


else:
    st.write("準備完了。サイドバーから設定を選び、[最適化を実行] を押してください。")
